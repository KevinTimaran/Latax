# ----------------------------------------------------------
# Fase 8.1: Exportar CSV final a Excel
# Proyecto: Localización de una fuente sísmica
# ----------------------------------------------------------

import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------
# 1. Rutas de entrada y salida
# ----------------------------------------------------------

ruta_csv = os.path.join("datos", "datos_sismicos_simulados_final.csv")
ruta_excel = os.path.join("datos", "datos_sismicos_simulados_final.xlsx")

# ----------------------------------------------------------
# 2. Leer datos del CSV final
# ----------------------------------------------------------

datos = []

with open(ruta_csv, mode="r", encoding="utf-8") as archivo_csv:
    lector = csv.reader(archivo_csv)
    for fila in lector:
        datos.append(fila)

encabezados = datos[0]
filas = datos[1:]

# ----------------------------------------------------------
# 3. Crear libro de Excel
# ----------------------------------------------------------

wb = Workbook()
ws = wb.active
ws.title = "Datos simulados"

# ----------------------------------------------------------
# 4. Título principal
# ----------------------------------------------------------

ws.merge_cells("A1:I1")
celda_titulo = ws["A1"]
celda_titulo.value = "Datos simulados de sensores sísmicos"
celda_titulo.font = Font(bold=True, size=14, color="FFFFFF")
celda_titulo.fill = PatternFill("solid", fgColor="1F4E78")
celda_titulo.alignment = Alignment(horizontal="center", vertical="center")

# ----------------------------------------------------------
# 5. Escribir encabezados
# ----------------------------------------------------------

fila_encabezados = 3

for col, encabezado in enumerate(encabezados, start=1):
    celda = ws.cell(row=fila_encabezados, column=col)
    celda.value = encabezado
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor="5B9BD5")
    celda.alignment = Alignment(horizontal="center", vertical="center")

# ----------------------------------------------------------
# 6. Escribir filas de datos
# ----------------------------------------------------------

for fila_idx, fila in enumerate(filas, start=4):
    for col_idx, valor in enumerate(fila, start=1):

        celda = ws.cell(row=fila_idx, column=col_idx)

        # Primera columna: Sensor
        if col_idx == 1:
            celda.value = valor
        else:
            celda.value = float(valor)

        celda.alignment = Alignment(horizontal="center", vertical="center")

# ----------------------------------------------------------
# 7. Aplicar bordes y formato numérico
# ----------------------------------------------------------

borde_fino = Side(style="thin", color="BFBFBF")
borde = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)

ultima_fila = 3 + len(filas)
ultima_columna = len(encabezados)

for fila in range(3, ultima_fila + 1):
    for columna in range(1, ultima_columna + 1):
        celda = ws.cell(row=fila, column=columna)
        celda.border = borde

        if columna >= 2:
            celda.number_format = "0.000000"

# Formato especial para coordenadas y distancia
for fila in range(4, ultima_fila + 1):
    ws.cell(row=fila, column=2).number_format = "0.0"
    ws.cell(row=fila, column=3).number_format = "0.0"
    ws.cell(row=fila, column=4).number_format = "0.0"
    ws.cell(row=fila, column=5).number_format = "0.0000"

# ----------------------------------------------------------
# 8. Ajustar ancho de columnas
# ----------------------------------------------------------

anchos = {
    "A": 12,
    "B": 10,
    "C": 10,
    "D": 10,
    "E": 12,
    "F": 16,
    "G": 14,
    "H": 14,
    "I": 18
}

for columna, ancho in anchos.items():
    ws.column_dimensions[columna].width = ancho

ws.row_dimensions[1].height = 25

# ----------------------------------------------------------
# 9. Congelar encabezados
# ----------------------------------------------------------

ws.freeze_panes = "A4"

# ----------------------------------------------------------
# 10. Crear hoja de resumen
# ----------------------------------------------------------

resumen = wb.create_sheet("Resumen")

resumen["A1"] = "Resumen de la simulación"
resumen["A1"].font = Font(bold=True, size=14, color="FFFFFF")
resumen["A1"].fill = PatternFill("solid", fgColor="1F4E78")
resumen.merge_cells("A1:B1")

resumen_datos = [
    ["Fuente real simulada", "(1.2, -0.8, -1.1)"],
    ["Amplitud inicial A0", "10.0"],
    ["Factor de ruido α", "0.05"],
    ["Número de sensores", "12"],
    ["Archivo base", "datos_sismicos_simulados_final.csv"],
    ["Archivo Excel generado", "datos_sismicos_simulados_final.xlsx"]
]

for fila_idx, fila in enumerate(resumen_datos, start=3):
    resumen.cell(row=fila_idx, column=1).value = fila[0]
    resumen.cell(row=fila_idx, column=2).value = fila[1]

    resumen.cell(row=fila_idx, column=1).font = Font(bold=True)
    resumen.cell(row=fila_idx, column=1).fill = PatternFill("solid", fgColor="D9EAF7")

    for col in range(1, 3):
        resumen.cell(row=fila_idx, column=col).border = borde
        resumen.cell(row=fila_idx, column=col).alignment = Alignment(horizontal="left")

resumen.column_dimensions["A"].width = 28
resumen.column_dimensions["B"].width = 38

# ----------------------------------------------------------
# 11. Crear gráfico de amplitud observada por sensor
# ----------------------------------------------------------

chart = BarChart()
chart.title = "Amplitud observada por sensor"
chart.y_axis.title = "Amplitud observada"
chart.x_axis.title = "Sensor"

datos_grafico = Reference(ws, min_col=9, min_row=3, max_row=ultima_fila)
categorias = Reference(ws, min_col=1, min_row=4, max_row=ultima_fila)

chart.add_data(datos_grafico, titles_from_data=True)
chart.set_categories(categorias)
chart.height = 8
chart.width = 16

ws.add_chart(chart, "K3")

# ----------------------------------------------------------
# 12. Guardar archivo Excel
# ----------------------------------------------------------

wb.save(ruta_excel)

print("----------------------------------------------------------")
print("Archivo Excel final generado correctamente")
print("----------------------------------------------------------")
print(f"Ruta del archivo: {ruta_excel}")
print(f"Total de registros exportados: {len(filas)}")
print("----------------------------------------------------------")